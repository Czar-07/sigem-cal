from __future__ import annotations

import hashlib
import os
import re
import zipfile

from datetime import date, datetime
from pathlib import Path
from tempfile import NamedTemporaryFile

from flask import current_app
from openpyxl import load_workbook

from app.database.database import db
from app.models.certificate import Certificate
from app.models.device import Device

from app.utils.r2_storage import (
    is_enabled as r2_enabled,
    list_certificates as r2_list_certificates,
    read_file as r2_read_file,
)


# ============================================================
# FILENAME PARSING
# ============================================================

FILENAME_RE = re.compile(
    r"""
    Certificado
    \s+
    de
    \s+
    calibra[cç][aã]o
    \s+
    (?P<num>\d+)
    \s*[-_]\s*
    (?P<year>20\d{2})
    \s*
    \(
    \s*
    DC
    \s*[-_\s]?\s*
    (?P<dc>[^)]+)
    \s*
    \)
    """,
    re.IGNORECASE | re.VERBOSE,
)


# ============================================================
# TEXTO
# ============================================================


def _text(value) -> str:
    """
    Converte qualquer valor para texto sem
    aplicar normalização de DC.
    """

    if value is None:
        return ""

    return str(value).strip()


# ============================================================
# NORMALIZAÇÃO DE DC
# ============================================================


def normalize_dc(value):
    """
    Normaliza o número do dispositivo.

    Exemplos:

        DC-737
        DC 737
        dc737
        0737
        737.0

    resultam em:

        737

    Identificadores especiais são preservados:

        DC-3.A
            -> 3.A

        DC-9.B
            -> 9.B

        DC-871_872
            -> 871_872

        DC-855_856
            -> 855_856

        DC-100000057300_400
            -> 100000057300_400
    """

    if value is None:
        return None

    value = str(value).strip().upper()

    if not value:
        return None

    # Remove extensão caso seja nome de arquivo.
    value = re.sub(
        r"\.(?:PDF|XLSX)$",
        "",
        value,
        flags=re.IGNORECASE,
    )

    # Remove prefixo DC.
    value = re.sub(
        r"^\s*DC[\s\-_]*",
        "",
        value,
        flags=re.IGNORECASE,
    )

    value = value.strip()

    if not value:
        return None

    # Excel pode transformar 737 em 737.0.
    if re.fullmatch(
        r"\d+\.0",
        value,
    ):
        value = value[:-2]

    # Remove zeros à esquerda somente
    # em identificadores numéricos puros.
    if re.fullmatch(
        r"\d+",
        value,
    ):
        value = str(
            int(value)
        )

    return value


def _normalize_dc(value):
    """
    Alias de compatibilidade.
    """

    normalized = normalize_dc(
        value
    )

    return normalized or ""


# ============================================================
# FILENAME METADATA
# ============================================================


def parse_filename(
    name: str,
):
    """
    Extrai informações do certificado pelo nome.

    Exemplos:

        Certificado de calibração 001-2026 (DC-918).pdf

        Certificado de calibração 168-2026 (DC-3.A).pdf

        Certificado de calibração 215-2026 (DC-871_872).pdf

        Certificado de calibração 046-2026
        (DC-100000057300_400).pdf
    """

    if not name:
        return None

    stem = Path(
        str(name)
    ).stem

    # --------------------------------------------------------
    # CERTIFICADO PADRÃO
    # --------------------------------------------------------

    match = FILENAME_RE.search(
        stem
    )

    if match:

        dc = normalize_dc(
            match.group("dc")
        )

        return {
            "numero_certificado": (
                f"{match.group('num')}/"
                f"{match.group('year')}"
            ),
            "ano": int(
                match.group("year")
            ),
            "dc": dc,
            "kind": "padrao",
        }

    # --------------------------------------------------------
    # TERCEIROS / RELATÓRIOS
    # --------------------------------------------------------

    dc_match = re.search(
        r"""
        (?:^|[/\\\s_(\-])
        DC
        [\s\-_]*
        (
            [A-Za-z0-9]
            [A-Za-z0-9._\-]*
        )
        """,
        str(name),
        re.IGNORECASE | re.VERBOSE,
    )

    year_match = re.search(
        r"(?:19|20)\d{2}",
        str(name),
    )

    if dc_match and year_match:

        dc = normalize_dc(
            dc_match.group(1)
        )

        return {
            "numero_certificado": stem,
            "ano": int(
                year_match.group(0)
            ),
            "dc": dc,
            "kind": "terceiros",
        }

    return None


# ============================================================
# DATE PARSING
# ============================================================


def _date_value(
    value,
) -> date | None:
    """
    Converte valores comuns do Excel para date.
    """

    if isinstance(
        value,
        datetime,
    ):
        return value.date()

    if isinstance(
        value,
        date,
    ):
        return value

    if not value:
        return None

    text = str(
        value
    ).strip()

    formats = (
        "%Y-%m-%d",
        "%d/%m/%Y",
        "%d-%m-%Y",
        "%Y/%m/%d",
        "%d.%m.%Y",
    )

    for fmt in formats:

        try:
            return datetime.strptime(
                text,
                fmt,
            ).date()

        except ValueError:
            continue

    return None


# ============================================================
# XLSX METADATA
# ============================================================


def read_xlsx_metadata(
    path: Path,
) -> dict:
    """
    Extrai metadados de certificados XLSX.

    A função é tolerante a diferentes
    modelos de certificado.
    """

    data = {}

    workbook = None

    try:

        workbook = load_workbook(
            path,
            data_only=True,
            read_only=True,
        )

        worksheet = (
            workbook.active
        )

        # ----------------------------------------------------
        # PRIMEIRAS 50 LINHAS
        # ----------------------------------------------------

        rows = list(
            worksheet.iter_rows(
                values_only=True
            )
        )

        for row in rows[:50]:

            values = list(row)

            for index, value in enumerate(
                values
            ):

                label = (
                    _text(value)
                    .lower()
                    .rstrip(":")
                )

                next_value = (
                    values[index + 1]
                    if index + 1
                    < len(values)
                    else None
                )

                # --------------------------------------------
                # DATA DE EMISSÃO / CALIBRAÇÃO
                # --------------------------------------------

                if (
                    "data de emissão"
                    in label
                    or "data da emissão"
                    in label
                    or "data de calibração"
                    in label
                    or "data da calibração"
                    in label
                ):

                    parsed = _date_value(
                        next_value
                    )

                    if parsed:
                        data[
                            "data_emissao"
                        ] = parsed

                # --------------------------------------------
                # LABORATÓRIO
                # --------------------------------------------

                if (
                    "laboratório"
                    in label
                    or "laboratorio"
                    in label
                ):

                    laboratory = _text(
                        next_value
                    )

                    if laboratory:
                        data[
                            "laboratorio"
                        ] = laboratory

                # --------------------------------------------
                # VALIDADE
                # --------------------------------------------

                if "validade" in label:

                    parsed = _date_value(
                        next_value
                    )

                    if parsed:

                        data[
                            "data_validade"
                        ] = parsed

            # ------------------------------------------------
            # REPRESENTAÇÃO TEXTUAL
            # ------------------------------------------------

            text = " | ".join(
                _text(value)
                for value in values
                if value is not None
            )

            validity_match = re.search(
                r"""
                Validade
                \s*:\s*
                (
                    \d{2}/
                    \d{2}/
                    \d{4}
                )
                """,
                text,
                re.IGNORECASE | re.VERBOSE,
            )

            if (
                validity_match
                and not data.get(
                    "data_validade"
                )
            ):

                data[
                    "data_validade"
                ] = _date_value(
                    validity_match.group(1)
                )

        # ----------------------------------------------------
        # RESULTADO
        # ----------------------------------------------------

        statuses = []

        for row in rows:

            for value in row:

                status = _text(
                    value
                ).upper()

                if status in {
                    "A",
                    "R",
                    "APROVADO",
                    "REPROVADO",
                }:

                    statuses.append(
                        status
                    )

        if any(
            status in {
                "R",
                "REPROVADO",
            }
            for status in statuses
        ):

            data[
                "resultado"
            ] = "REPROVADO"

        elif any(
            status in {
                "A",
                "APROVADO",
            }
            for status in statuses
        ):

            data[
                "resultado"
            ] = "APROVADO"

    except Exception:
        # A falha de leitura de metadata
        # nunca deve impedir a sincronização.
        pass

    finally:

        if workbook is not None:

            try:
                workbook.close()
            except Exception:
                pass

    return data


# ============================================================
# FINGERPRINT
# ============================================================


def fingerprint_bytes(
    data: bytes,
) -> str:
    """
    SHA-256 dos bytes.
    """

    return hashlib.sha256(
        data
    ).hexdigest()


def fingerprint_file(
    path: Path,
) -> str:
    """
    SHA-256 de arquivo.
    """

    digest = hashlib.sha256()

    with path.open(
        "rb"
    ) as file:

        for chunk in iter(
            lambda: file.read(
                1024 * 1024
            ),
            b"",
        ):

            digest.update(
                chunk
            )

    return digest.hexdigest()


# ============================================================
# R2 SOURCE
# ============================================================


def r2_source_candidates():
    """
    Descobre certificados no Cloudflare R2.

    Prefixo padrão:

        Certificados/

    Exemplo de chave:

        Certificados/Certificados de Calibração 2026/
        Certificado de calibração 001-2026 (DC-918).xlsx
    """

    if not r2_enabled():
        return

    for item in r2_list_certificates():

        yield {
            "key": item["key"],
            "name": item["name"],
            "ext": item["ext"],
            "r2": True,
            "size": item.get(
                "size",
                0,
            ),
            "etag": item.get(
                "etag",
                "",
            ),
            "signature": (
                f'{item.get("etag", "")}:'
                f'{item.get("size", 0)}'
            ),
        }


# ============================================================
# SOURCE DISCOVERY
# ============================================================


def source_candidates(
    source: Path,
):
    """
    Descobre certificados em:

        diretório

    ou:

        ZIP

    Suportados:

        .pdf
        .xlsx
    """

    # --------------------------------------------------------
    # DIRECTORY
    # --------------------------------------------------------

    if source.is_dir():

        for path in source.rglob("*"):

            if not path.is_file():
                continue

            if path.name.startswith(
                "~$"
            ):
                continue

            if path.suffix.lower() not in {
                ".pdf",
                ".xlsx",
            }:
                continue

            try:
                stat = path.stat()
            except OSError:
                continue

            yield {
                "key": path.relative_to(
                    source
                ).as_posix(),

                "name": path.name,

                "ext": path.suffix.lower(),

                "path": path,

                "signature": (
                    f"{stat.st_size}:"
                    f"{stat.st_mtime_ns}"
                ),
            }

        return

    # --------------------------------------------------------
    # ZIP
    # --------------------------------------------------------

    if (
        source.is_file()
        and source.suffix.lower()
        == ".zip"
    ):

        try:

            with zipfile.ZipFile(
                source
            ) as archive:

                for info in archive.infolist():

                    name = (
                        info.filename
                        .replace(
                            "\\",
                            "/",
                        )
                    )

                    filename = Path(
                        name
                    ).name

                    extension = Path(
                        name
                    ).suffix.lower()

                    if info.is_dir():
                        continue

                    if filename.startswith(
                        "~$"
                    ):
                        continue

                    if extension not in {
                        ".pdf",
                        ".xlsx",
                    }:
                        continue

                    signature = hashlib.sha256(
                        (
                            f"{info.CRC}:"
                            f"{info.file_size}:"
                            f"{info.date_time}"
                        ).encode(
                            "utf-8"
                        )
                    ).hexdigest()

                    yield {
                        "key": name,
                        "name": filename,
                        "ext": extension,
                        "zip": source,
                        "info": info,
                        "signature": signature,
                    }

        except zipfile.BadZipFile:
            return


# ============================================================
# SOURCE READING
# ============================================================


def _read_source(
    candidate: dict,
) -> bytes:
    """
    Lê certificado de:

        R2
        diretório local
        ZIP
    """

    if candidate.get("r2"):
        return r2_read_file(
            candidate["key"]
        )

    if "path" in candidate:
        return candidate[
            "path"
        ].read_bytes()

    with zipfile.ZipFile(
        candidate["zip"]
    ) as archive:

        return archive.read(
            candidate["info"]
        )


def read_source_member(
    source_path: str,
    source_key: str,
):
    """
    Lê diretamente um membro do ZIP original.

    Retorna:

        (bytes, extensão)

    ou:

        None
    """

    if not source_path or not source_key:
        return None

    try:

        source = (
            Path(
                source_path
            )
            .expanduser()
            .resolve()
        )

    except OSError:
        return None

    if (
        not source.is_file()
        or source.suffix.lower()
        != ".zip"
    ):
        return None

    try:

        with zipfile.ZipFile(
            source
        ) as archive:

            normalized_key = (
                source_key
                .replace(
                    "\\",
                    "/",
                )
            )

            info = archive.getinfo(
                normalized_key
            )

            data = archive.read(
                info
            )

            extension = Path(
                normalized_key
            ).suffix.lower()

            return (
                data,
                extension,
            )

    except (
        KeyError,
        OSError,
        zipfile.BadZipFile,
    ):
        return None


# ============================================================
# UPLOAD STORAGE LOCAL
# ============================================================


def _upload_root() -> Path:
    """
    Retorna o diretório local de uploads.

    Usado somente quando a fonte
    NÃO é R2.
    """

    configured = current_app.config.get(
        "UPLOAD_FOLDER"
    )

    if not configured:

        configured = (
            Path(
                current_app.instance_path
            )
            / "uploads"
        )

    root = Path(
        configured
    ).resolve()

    root.mkdir(
        parents=True,
        exist_ok=True,
    )

    return root


def _ensure_file(
    data: bytes,
    ano: int,
    filename: str,
) -> str:
    """
    Armazena certificado localmente.

    Estrutura:

        uploads/
            certificates/
                2026/
                    certificado.pdf
    """

    upload_root = _upload_root()

    base = (
        upload_root
        / "certificates"
        / str(ano)
    )

    base.mkdir(
        parents=True,
        exist_ok=True,
    )

    safe_filename = re.sub(
        r"[^A-Za-z0-9À-ÿ.\-_() ]+",
        "_",
        filename,
    ).strip(
        " ._"
    )

    if not safe_filename:
        safe_filename = (
            "certificado.pdf"
        )

    target = (
        base
        / safe_filename
    )

    if (
        not target.exists()
        or fingerprint_file(target)
        != fingerprint_bytes(data)
    ):

        target.write_bytes(
            data
        )

    return target.relative_to(
        upload_root
    ).as_posix()


# ============================================================
# DISCOVER SOURCE
# ============================================================


def _discover_source(
    source_value=None,
):
    """
    Descobre fonte local.

    Ordem:

        1. fonte explicitamente informada
        2. CERTIFICATES_FOLDER
        3. Certificados.zip
        4. certificados.zip
        5. Certificados/
        6. certificados/

    IMPORTANTE:

        R2 é tratado separadamente por
        synchronize_certificates().
    """

    candidates = []

    # --------------------------------------------------------
    # FONTE EXPLÍCITA
    # --------------------------------------------------------

    if source_value:

        raw = str(
            source_value
        ).strip()

        if raw:

            candidates.append(
                Path(
                    raw
                ).expanduser()
            )

    # --------------------------------------------------------
    # DIRETÓRIO DO PROJETO
    # --------------------------------------------------------

    base = (
        Path(__file__)
        .resolve()
        .parents[2]
    )

    cwd = Path.cwd().resolve()

    # --------------------------------------------------------
    # CONFIGURAÇÃO
    # --------------------------------------------------------

    configured = current_app.config.get(
        "CERTIFICATES_FOLDER"
    )

    if configured:

        candidates.insert(
            0,
            Path(
                configured
            ).expanduser(),
        )

    # --------------------------------------------------------
    # CANDIDATOS
    # --------------------------------------------------------

    candidates.extend(
        [
            base / "Certificados.zip",
            base / "certificados.zip",
            base / "Certificados",
            base / "certificados",

            base.parent / "Certificados.zip",
            base.parent / "certificados.zip",

            cwd / "Certificados.zip",
            cwd / "certificados.zip",
            cwd / "Certificados",
            cwd / "certificados",
        ]
    )

    seen = set()

    for candidate in candidates:

        try:
            resolved = candidate.resolve()
        except OSError:
            continue

        normalized = str(
            resolved
        ).lower()

        if normalized in seen:
            continue

        seen.add(
            normalized
        )

        if not resolved.exists():
            continue

        if resolved.is_dir():
            return resolved

        if (
            resolved.is_file()
            and resolved.suffix.lower()
            == ".zip"
        ):
            return resolved

    return None


# ============================================================
# DEVICE LOOKUP
# ============================================================


def _find_device_by_dc(
    dc,
):
    """
    Localiza Device através do DC normalizado.

    Exemplo:

        Certificado:
            DC-918

        Device:
            DC-918

    Ambos:

        918
    """

    normalized_dc = normalize_dc(
        dc
    )

    if not normalized_dc:
        return None

    devices = Device.query.all()

    for device in devices:

        device_dc = normalize_dc(
            getattr(
                device,
                "numero",
                None,
            )
        )

        if device_dc == normalized_dc:
            return device

    return None


# ============================================================
# CERTIFICATE GROUPING
# ============================================================


def _group_candidates(
    candidates,
    stats,
):
    """
    Agrupa certificados logicamente.

    Para certificados padrão:

        ano
        número
        DC

    definem o certificado lógico.

    Se existir:

        XLSX
        PDF

    para o mesmo certificado:

        PDF ganha.
    """

    grouped = {}

    for candidate in candidates:

        meta = parse_filename(
            candidate["name"]
        )

        if not meta:

            stats[
                "ignored"
            ].append(
                candidate["name"]
            )

            continue

        candidate = {
            **candidate,
            "meta": meta,
        }

        # ----------------------------------------------------
        # PADRÃO
        # ----------------------------------------------------

        if meta["kind"] == "padrao":

            logical_key = (
                f"{meta['ano']}|"
                f"{meta['numero_certificado']}|"
                f"{normalize_dc(meta['dc'])}"
            ).upper()

            existing = grouped.get(
                logical_key
            )

            if existing is None:

                grouped[
                    logical_key
                ] = candidate

            else:

                # PDF sempre vence XLSX.
                if (
                    candidate["ext"]
                    == ".pdf"
                    and existing["ext"]
                    != ".pdf"
                ):

                    grouped[
                        logical_key
                    ] = candidate

        # ----------------------------------------------------
        # TERCEIROS
        # ----------------------------------------------------

        else:

            grouped[
                candidate["key"]
            ] = candidate

    return list(
        grouped.values()
    )


# ============================================================
# DATABASE DEVICE CACHE
# ============================================================


def _build_device_cache():
    """
    Cria um índice:

        DC normalizado -> Device
    """

    devices_by_dc = {}

    for device in Device.query.all():

        normalized = normalize_dc(
            getattr(
                device,
                "numero",
                None,
            )
        )

        if not normalized:
            continue

        # Não sobrescrever silenciosamente
        # um dispositivo já existente.
        if normalized not in devices_by_dc:

            devices_by_dc[
                normalized
            ] = device

    return devices_by_dc


# ============================================================
# MAIN SYNCHRONIZATION
# ============================================================


def synchronize_certificates(
    source_value=None,
):
    """
    Sincroniza certificados.

    Com R2 configurado:

        R2 é a fonte padrão.

    Com source_value informado:

        usa a fonte local explicitamente.

    Sem R2:

        procura automaticamente:

            Certificados.zip
            certificados.zip
            Certificados/
            certificados/
    """

    # ========================================================
    # DEFINIR FONTE
    # ========================================================

    explicit_source = bool(
        str(
            source_value or ""
        ).strip()
    )

    use_r2 = (
        r2_enabled()
        and not explicit_source
    )

    if use_r2:

        source = None

    else:

        source = _discover_source(
            source_value
        )

        if source is None:

            return {
                "success": False,

                "message": (
                    "Nenhuma fonte de "
                    "certificados foi encontrada."
                ),

                "source": None,

                "imported": 0,
                "updated": 0,
                "unchanged": 0,
                "scanned": 0,

                "unmatched": [],
                "ignored": [],
                "errors": [],
            }

    # ========================================================
    # ESTATÍSTICAS
    # ========================================================

    stats = {

        "success": True,

        "source": (
            "Cloudflare R2"
            if use_r2
            else str(source)
        ),

        "scanned": 0,
        "imported": 0,
        "updated": 0,
        "unchanged": 0,

        "unmatched": [],
        "ignored": [],
        "errors": [],
    }

    # ========================================================
    # DESCOBRIR CERTIFICADOS
    # ========================================================

    try:

        if use_r2:

            candidates = list(
                r2_source_candidates()
            )

        else:

            candidates = list(
                source_candidates(
                    source
                )
            )

        devices_by_dc = (
            _build_device_cache()
        )

    except Exception as exc:

        stats["success"] = False

        stats[
            "errors"
        ].append(
            {
                "arquivo": "__source__",
                "erro": str(exc),
            }
        )

        return stats

    # ========================================================
    # AGRUPAR
    # ========================================================

    grouped_candidates = (
        _group_candidates(
            candidates,
            stats,
        )
    )

    # ========================================================
    # PROCESSAR
    # ========================================================

    for candidate in grouped_candidates:

        stats[
            "scanned"
        ] += 1

        meta = candidate[
            "meta"
        ]

        try:

            # ------------------------------------------------
            # LER ARQUIVO
            # ------------------------------------------------

            raw = _read_source(
                candidate
            )

            fingerprint = (
                fingerprint_bytes(
                    raw
                )
            )

            # ------------------------------------------------
            # LOCALIZAR DEVICE PRIMEIRO
            # ------------------------------------------------

            normalized_dc = normalize_dc(
                meta.get("dc")
            )

            device = devices_by_dc.get(
                normalized_dc
            )

            if device is None:

                stats[
                    "unmatched"
                ].append(
                    {
                        "dc": meta.get(
                            "dc"
                        ),

                        "dc_normalizado":
                            normalized_dc,

                        "arquivo":
                            candidate[
                                "name"
                            ],

                        "source_key":
                            candidate[
                                "key"
                            ],
                    }
                )

                continue

            # ------------------------------------------------
            # PROCURAR CERTIFICADO EXISTENTE
            # ------------------------------------------------

            with db.session.no_autoflush:

                certificate = (
                    Certificate.query
                    .filter_by(
                        source_key=(
                            candidate[
                                "key"
                            ]
                        )
                    )
                    .first()
                )

            is_new = (
                certificate is None
            )

            # ------------------------------------------------
            # CERTIFICADO EXISTENTE SEM ALTERAÇÃO
            # ------------------------------------------------

            if (
                certificate is not None
                and certificate.source_hash
                == fingerprint
            ):

                changed_relation = (
                    certificate.device_id
                    != device.id
                )

                if changed_relation:

                    certificate.device_id = (
                        device.id
                    )

                    certificate.updated_at = (
                        datetime.utcnow()
                    )

                    stats[
                        "updated"
                    ] += 1

                else:

                    stats[
                        "unchanged"
                    ] += 1

                continue

            # ------------------------------------------------
            # NOVO / EXISTENTE
            # ------------------------------------------------

            if is_new:

                certificate = (
                    Certificate()
                )

            else:

                stats[
                    "updated"
                ] += 1

            # ------------------------------------------------
            # METADATA XLSX
            # ------------------------------------------------

            xlsx_meta = {}

            if candidate["ext"] == ".xlsx":

                temporary_path = None

                try:

                    with NamedTemporaryFile(
                        suffix=".xlsx",
                        delete=False,
                    ) as temporary_file:

                        temporary_file.write(
                            raw
                        )

                        temporary_path = Path(
                            temporary_file.name
                        )

                    xlsx_meta = (
                        read_xlsx_metadata(
                            temporary_path
                        )
                    )

                finally:

                    if temporary_path:

                        temporary_path.unlink(
                            missing_ok=True
                        )

            # ------------------------------------------------
            # PREENCHER CERTIFICATE
            # ------------------------------------------------

            certificate.device_id = (
                device.id
            )

            certificate.ano = (
                meta["ano"]
            )

            certificate.numero_certificado = (
                meta[
                    "numero_certificado"
                ]
            )

            certificate.nome_arquivo = (
                candidate["name"]
            )

            certificate.source_key = (
                candidate["key"]
            )

            certificate.source_hash = (
                fingerprint
            )

            certificate.source_type = (
                candidate["ext"]
                .lstrip(".")
                .lower()
            )

            # ------------------------------------------------
            # R2
            # ------------------------------------------------

            if candidate.get("r2"):

                bucket = os.getenv(
                    "R2_BUCKET",
                    "sigem-certificados",
                ).strip()

                certificate.source_path = (
                    f"r2://{bucket}"
                )

                certificate.arquivo = (
                    f"r2://{bucket}/"
                    f"{candidate['key']}"
                )

            # ------------------------------------------------
            # LOCAL
            # ------------------------------------------------

            else:

                certificate.source_path = (
                    str(source)
                )

                certificate.arquivo = (
                    _ensure_file(
                        raw,
                        meta["ano"],
                        candidate["name"],
                    )
                )

            # ------------------------------------------------
            # METADATA
            # ------------------------------------------------

            certificate.data_emissao = (
                xlsx_meta.get(
                    "data_emissao"
                )
            )

            certificate.data_validade = (
                xlsx_meta.get(
                    "data_validade"
                )
            )

            certificate.laboratorio = (
                xlsx_meta.get(
                    "laboratorio"
                )
            )

            certificate.resultado = (
                xlsx_meta.get(
                    "resultado"
                )
            )

            # ------------------------------------------------
            # OBSERVAÇÕES
            # ------------------------------------------------

            certificate.observacoes = (
                "Sincronizado automaticamente. "
                f"DC identificado pelo nome "
                f"do arquivo: {meta.get('dc')}."
            )

            certificate.updated_at = (
                datetime.utcnow()
            )

            # ------------------------------------------------
            # INSERT
            # ------------------------------------------------

            if is_new:

                db.session.add(
                    certificate
                )

                stats[
                    "imported"
                ] += 1

        except Exception as exc:

            db.session.rollback()

            stats[
                "errors"
            ].append(
                {
                    "arquivo": candidate[
                        "name"
                    ],

                    "source_key": candidate[
                        "key"
                    ],

                    "erro": str(exc),
                }
            )

    # ========================================================
    # COMMIT
    # ========================================================

    try:

        db.session.commit()

    except Exception as exc:

        db.session.rollback()

        stats["success"] = False

        stats[
            "errors"
        ].append(
            {
                "arquivo": "__commit__",
                "erro": str(exc),
            }
        )

    # ========================================================
    # MENSAGEM FINAL
    # ========================================================

    if stats["success"]:

        stats[
            "message"
        ] = (
            "Sincronização concluída."
        )

    else:

        stats[
            "message"
        ] = (
            "Sincronização concluída "
            "com erros."
        )

    return stats